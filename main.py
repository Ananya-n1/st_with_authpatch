import streamlit as st
import yaml
import bcrypt
import streamlit_authenticator as stauth
import pandas as pd
from openpyxl import load_workbook
import os



if 'in_session' not in st.session_state:
    st.session_state.in_session = False

def landing():
    st.title("Streamlit application")

    option = st.radio('Choose an Operation: ', ['Login', 'Signup'])
    if option == 'Login':
        login()

    elif option == 'Signup':
        signup()

def load_credentials():
    with open('config.yaml', 'r') as file:
        config = yaml.safe_load(file)
    return config


def save_credentials(new_user):
    with open('config.yaml', 'r') as file:
        config = yaml.safe_load(file)

    config['credentials']['usernames'].update(new_user)

    with open('config.yaml', 'w') as file:
        yaml.dump(config, file)


def save_data_to_excel(name, choice):
    file_name = 'form_data.xlsx'
    data = pd.DataFrame([[name, choice]], columns=['Name', 'Gender'])

    if not os.path.exists(file_name):
        # If the file doesn't exist, create it
        data.to_excel(file_name, index=False)
    else:
        try:

            existing_data = pd.read_excel(file_name)
            duplicate = (existing_data['Name'] == name) & (existing_data['Gender'] == choice)

            if duplicate.any():
                st.warning(f"Duplicate entry found: {name}, {choice}")
            else:
                with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    # Getting the last row in the existing sheet
                    book = load_workbook(file_name)
                    sheet = book.active
                    startrow = sheet.max_row

                    data.to_excel(writer, index=False, header=False, startrow=startrow)
                    st.success(f"Data saved: {name}, {choice}")
        except Exception as e:
            st.error(f"Error: {e}")

def login():
    st.header("Login")

    config = load_credentials()
    authenticator = stauth.Authenticate(
        config['credentials'],
        config['cookie']['name'],
        config['cookie']['key'],
        config['cookie']['expiry_days']
    )

    name, auth_status, email = authenticator.login("main",10,3)

    if auth_status:
        user_role = config['credentials']['usernames'][email]['role']
        if user_role == 'admin':
            admin_dashboard()
            # st.write('<meta http-equiv="refresh" content="0; url=/admin_dashboard" />', unsafe_allow_html=True)
        elif user_role == "user":
            dashboard()
            # st.write('<meta http-equiv="refresh" content="0; url=/dashboard" />', unsafe_allow_html=True)

    elif auth_status is False:
        st.error("ERROR: Invalid Credentials. Please try again.")

    elif auth_status is None:
        st.warning("Please enter your email and password.")

def admin_dashboard():
    #if st.session_state['in_session']:
        st.success("Login successful!, Welcome admin,")
        st.title('Data Collection Form')

        with st.form(key='data_form'):
            name = st.text_input('Enter your name')
            choice = st.radio('Gender:', ['Male', 'Female', 'Other'])

            submit_button = st.form_submit_button(label='Submit')

        if st.button('Load Data'):
            filename = 'form_data.xlsx'
            if os.path.exists(filename):
                df = pd.read_excel(filename)
                st.dataframe(df)
            else:
                st.warning("No data available to show, fill the form first")

        if submit_button:
            if name and choice:
                save_data_to_excel(name, choice)
            else:
                st.error("Please enter your name.")


def dashboard():

   # if cr.key == 1:
        st.success("Login successful! Welcome")
        st.header("You only have reading permissions")
        if st.button('Load Data'):
            filename = 'form_data.xlsx'
            if os.path.exists(filename):
                df = pd.read_excel(filename)
                st.dataframe(df)
            else:
                st.warning("No data available to show, fill the form first")



def signup():
    st.subheader("Sign Up")
    name = st.text_input("Your  Username")
    email = st.text_input("Email")
    password = st.text_input("Password", type="password")

    try:
        if st.button("Sign Up"):

            if email and password and name:
                users = load_credentials()
                if name in users['credentials']['usernames']:
                    st.error("Username already exists!")
                else:
                    salt = bcrypt.gensalt()
                    hashed_pw = bcrypt.hashpw(password.encode('utf-8'), salt)
                    new_user = {
                        name: {
                            'mail': email,
                            'password': hashed_pw.decode('utf-8'),
                            'role': 'user'
                        }
                    }
                    save_credentials(new_user)
                    st.success("Registration Successful redirecting to login")

            else:
                st.error("Please enter all the details")


    except Exception as e:
        st.error(f"Error: An error occured: {e}")

if __name__ == "__main__":
    landing()